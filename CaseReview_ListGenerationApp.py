import os
import re
import time
import requests
import pandas as pd
from datetime import datetime
from urllib.parse import quote
import msal
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

# ==============================
# Config (billing cycle label only; hours not used here)
# ==============================
CYCLE_START_LABEL = "09/24/25"
CYCLE_END_LABEL   = "10/07/25"
BILLING_COL = f"Billing Cycle Hours ({CYCLE_START_LABEL} - {CYCLE_END_LABEL})"

# ==============================
# Environment / Clio / Graph
# ==============================
API_VERSION = "4"
CLIO_BASE = os.getenv("CLIO_BASE", "https://app.clio.com").rstrip("/")
CLIO_API = f"{CLIO_BASE}/api/v{API_VERSION}"
CLIO_TOKEN_URL = f"{CLIO_BASE}/oauth/token"

CLIO_CLIENT_ID = os.getenv("CLIO_CLIENT_ID")
CLIO_CLIENT_SECRET = os.getenv("CLIO_CLIENT_SECRET")
CLIO_REFRESH_TOKEN = os.getenv("CLIO_REFRESH_TOKEN")
_CLIO_ACCESS_TOKEN = os.getenv("CLIO_ACCESS_TOKEN") or ""
_CLIO_EXPIRES_AT_ENV = os.getenv("CLIO_EXPIRES_AT") or os.getenv("CLIO_EXPIRES_IN")
try:
    _CLIO_EXPIRES_AT = float(_CLIO_EXPIRES_AT_ENV) if _CLIO_EXPIRES_AT_ENV else 0.0
except Exception:
    _CLIO_EXPIRES_AT = 0.0

SHAREPOINT_TENANT_ID = os.getenv("GRAPH_TENANT_ID")
SHAREPOINT_CLIENT_ID = os.getenv("GRAPH_CLIENT_ID")
SHAREPOINT_CLIENT_SECRET = os.getenv("GRAPH_CLIENT_SECRET")
SHAREPOINT_SITE_ID = os.getenv("SHAREPOINT_SITE_ID")
SHAREPOINT_DRIVE_ID = os.getenv("SHAREPOINT_DRIVE_ID")
SHAREPOINT_DOC_LIB = os.getenv("SHAREPOINT_DOC_LIB", "General Management/Global Case Review Lists")

# ==============================
# Attorney folder mapping (by last name)
# Use these for exceptions or custom folder names.
# Everyone else will be handled automatically by the smart splitter.
# ==============================
ATTORNEY_FOLDERS = {
    "voorhees": "Attorneys and Paralegals/Attorney Case Lists/Voorhees, Elizabeth",
    "darling":  "Attorneys and Paralegals/Attorney Case Lists/Darling, Craig",
    "huang":    "Attorneys and Paralegals/Attorney Case Lists/Huang, Lily",
    "parker":   "Attorneys and Paralegals/Attorney Case Lists/Parker, Gabriella",
    # Optional: add overrides here if needed, e.g.:
    # "kirsten": "Attorneys and Paralegals/Attorney Case Lists/Kirsten, Natalie",
}
ATTORNEY_DISPLAY = {
    "voorhees": "Voorhees, Elizabeth",
    "darling":  "Darling, Craig",
    "huang":    "Huang, Lily",
    "parker":   "Parker, Gabriella",
    # "kirsten":  "Kirsten, Natalie",
}

# Base for automatically derived folders
BASE_ATTY_FOLDER = "Attorneys and Paralegals/Attorney Case Lists"

# ==============================
# Output columns
# ==============================
OUTPUT_FIELDS = [
    "Matter Number","Client Name","CR ID","Net Trust Account Balance","Matter Stage",
    BILLING_COL,
    "Responsible Attorney","Main Paralegal","Supporting Attorney","Supporting Paralegal","Client Notes",
    "Initial Client Goals","Initial Strategy","Has strategy changed Describe","Current action Items",
    "Hearings","Deadlines","DV situation description","Custody Visitation","CS Add ons Extracurricular",
    "Spousal Support","PDDs","Discovery","Judgment Trial","Post Judgment","collection efforts"
]

# ==============================
# HTTP helpers
# ==============================
PAGE_LIMIT = 200
GLOBAL_MIN_SLEEP = float(os.getenv("CLIO_GLOBAL_MIN_SLEEP_SEC", "0.25"))
session = requests.Session()
session.headers.update({"Accept": "application/json"})

def _save_tokens_env(tokens: dict):
    global _CLIO_ACCESS_TOKEN, _CLIO_EXPIRES_AT, CLIO_REFRESH_TOKEN
    _CLIO_ACCESS_TOKEN = tokens.get("access_token", _CLIO_ACCESS_TOKEN)
    if tokens.get("refresh_token"):
        CLIO_REFRESH_TOKEN = tokens["refresh_token"]
        os.environ["CLIO_REFRESH_TOKEN"] = CLIO_REFRESH_TOKEN
    expires_in = tokens.get("expires_in", 0)
    try:
        _CLIO_EXPIRES_AT = time.time() + float(expires_in)
    except Exception:
        _CLIO_EXPIRES_AT = time.time() + 3000
    os.environ["CLIO_ACCESS_TOKEN"] = _CLIO_ACCESS_TOKEN or ""
    os.environ["CLIO_EXPIRES_AT"] = str(_CLIO_EXPIRES_AT)

def _refresh_clio_token() -> str:
    if not (CLIO_CLIENT_ID and CLIO_CLIENT_SECRET and CLIO_REFRESH_TOKEN):
        raise RuntimeError("Missing CLIO credentials.")
    resp = session.post(CLIO_TOKEN_URL, data={
        "grant_type": "refresh_token",
        "refresh_token": CLIO_REFRESH_TOKEN,
        "client_id": CLIO_CLIENT_ID,
        "client_secret": CLIO_CLIENT_SECRET
    }, timeout=60)
    if resp.status_code != 200:
        raise RuntimeError(f"Token refresh failed: {resp.status_code} {resp.text[:200]}")
    tokens = resp.json() or {}
    _save_tokens_env(tokens)
    return _CLIO_ACCESS_TOKEN

def _get_clio_token() -> str:
    now = time.time()
    if (not _CLIO_ACCESS_TOKEN) or now >= (_CLIO_EXPIRES_AT or 0):
        return _refresh_clio_token()
    return _CLIO_ACCESS_TOKEN

def _ensure_auth_header():
    token = _get_clio_token()
    session.headers["Authorization"] = f"Bearer {token}"

def _sleep_with_floor(start_ts: float, retry_after: int | float | None = None):
    elapsed = time.time() - start_ts
    base_wait = max(0, GLOBAL_MIN_SLEEP - elapsed)
    wait = max(base_wait, float(retry_after or 0))
    if wait > 0:
        time.sleep(wait)

def _request(method: str, url: str, **kwargs) -> requests.Response:
    _ensure_auth_header()
    max_tries = kwargs.pop("max_tries", 7)
    backoff = 1
    for _ in range(max_tries):
        t0 = time.time()
        resp = session.request(method, url, timeout=60, **kwargs)
        if resp.status_code == 401:
            _refresh_clio_token()
            _sleep_with_floor(t0, retry_after=1)
            continue
        if resp.status_code == 429:
            ra = 30
            if resp.headers.get("Retry-After"):
                try: ra = int(resp.headers["Retry-After"])
                except: ra = 30
            print(f"[429] Waiting {ra}s…")
            _sleep_with_floor(t0, retry_after=ra)
            continue
        if 500 <= resp.status_code < 600:
            print(f"[{resp.status_code}] Server error. Retrying in {backoff}s…")
            _sleep_with_floor(t0, retry_after=backoff)
            backoff = min(backoff * 2, 60)
            continue
        _sleep_with_floor(t0)
        return resp
    return resp

def paginate(url: str, params: dict | None = None):
    params = dict(params or {})
    params.setdefault("limit", PAGE_LIMIT)
    params.setdefault("order", "id(asc)")
    next_url, next_params = url, params
    all_rows = []
    while True:
        resp = _request("GET", next_url, params=next_params)
        print(f"GET {next_url} params={next_params} → {resp.status_code}")
        if resp.status_code != 200:
            print(f"Failed: {resp.status_code} {resp.text[:200]}")
            break
        body = resp.json() or {}
        rows = body.get("data", [])
        if isinstance(rows, list):
            all_rows.extend([r for r in rows if isinstance(r, dict)])
        paging = (body.get("meta") or {}).get("paging") or {}
        if paging.get("next"):
            next_url, next_params = paging["next"], None
            continue
        break
    return all_rows

# ==============================
# Utils
# ==============================
def _num(x):
    try:
        return float(x or 0)
    except Exception:
        return 0.0

def _normalize_name(name: str) -> str:
    name = (name or "").strip()
    if not name:
        return ""
    if "," in name:
        last, first = name.split(",", 1)
        return f"{first.strip()} {last.strip()}"
    return name

def _to_last_first(name: str) -> str:
    """
    Convert 'First Middle Last' or 'Last, First Middle' to 'Last, First Middle'
    """
    name = (name or "").strip()
    if not name:
        return ""
    if "," in name:
        last, first = [p.strip() for p in name.split(",", 1)]
        return f"{last}, {first}"
    parts = name.split()
    if len(parts) == 1:
        return parts[0]
    first = " ".join(parts[:-1])
    last = parts[-1]
    return f"{last}, {first}"

# ==============================
# Clio fetchers
# ==============================
def fetch_custom_fields_meta():
    rows = paginate(f"{CLIO_API}/custom_fields.json", {"fields": "id,name,field_type,picklist_options"})
    out = {}
    for f in rows:
        name = f.get("name")
        if not name:
            continue
        opts = {str(o["id"]): o["option"] for o in f.get("picklist_options", []) if o.get("id")}
        out[name] = {"id": f.get("id"), "type": f.get("field_type"), "options": opts}
    return out

def fetch_open_matters_with_cf():
    fields = ("id,display_number,number,client{id,name},matter_stage{name},"
              "responsible_attorney{name},account_balances{balance},"
              "custom_field_values{id,field_name,field_type,value,picklist_option}")
    return paginate(f"{CLIO_API}/matters.json", {"status": "open,pending", "fields": fields})

def fetch_outstanding_client_balances():
    rows = paginate(f"{CLIO_API}/outstanding_client_balances.json",
                    {"fields": "contact{id,name},total_outstanding_balance"})
    return [{
        "Client Name": _normalize_name((r.get("contact") or {}).get("name") or ""),
        "Outstanding Balance": _num(r.get("total_outstanding_balance"))
    } for r in rows]

def fetch_billable_matters_client_unbilled():
    rows = paginate(f"{CLIO_API}/billable_matters.json",
                    {"fields": "id,display_number,client{name},unbilled_amount,unbilled_hours"})
    agg: dict[str, dict] = {}
    for r in rows:
        client_name = _normalize_name(((r.get("client") or {}).get("name")) or "")
        a = agg.setdefault(client_name, {"Unbilled Amount": 0.0, "Unbilled Hours": 0.0})
        a["Unbilled Amount"] += _num(r.get("unbilled_amount"))
        a["Unbilled Hours"] += _num(r.get("unbilled_hours"))
    out = []
    for name, v in agg.items():
        out.append({
            "Client Name": name,
            "Unbilled Amount": v["Unbilled Amount"],
            "Unbilled Hours": v["Unbilled Hours"],
        })
    return out

# ==============================
# Report builder
# ==============================
CF_OUTPUT_FIELDS = [
    "CR ID","Main Paralegal","Supporting Attorney","Supporting Paralegal","Client Notes",
    "Initial Client Goals","Initial Strategy","Has strategy changed Describe","Current action Items",
    "Hearings","Deadlines","DV situation description","Custody Visitation","CS Add ons Extracurricular",
    "Spousal Support","PDDs","Discovery","Judgment Trial","Post Judgment","collection efforts"
]

def _resolve_cf_value(cf: dict, meta_by_name: dict) -> str:
    if cf.get("field_type") == "picklist":
        opt = (cf.get("picklist_option") or {}).get("option")
        if opt:
            return opt
        raw = cf.get("value")
        options = ((meta_by_name.get(cf.get("field_name")) or {}).get("options")) or {}
        return options.get(str(raw), raw) or ""
    return str(cf.get("value") or "")

def build_report_dataframe() -> pd.DataFrame:
    cf_meta = fetch_custom_fields_meta()
    matters = fetch_open_matters_with_cf()
    ocb_rows = fetch_outstanding_client_balances()
    unbilled_rows = fetch_billable_matters_client_unbilled()

    # -------- Matters (per matter) --------
    m_rows = []
    for m in matters:
        trust = 0.0
        for b in (m.get("account_balances") or []):
            trust += _num((b or {}).get("balance"))
        cf_map = {}
        for cf in (m.get("custom_field_values") or []):
            name = cf.get("field_name")
            if not name:
                continue
            cf_map[name] = _resolve_cf_value(cf, cf_meta)
        m_rows.append({
            "Matter ID": m.get("id"),
            "Matter Number": m.get("display_number") or m.get("number") or "",
            "Client Name": _normalize_name(((m.get("client") or {}).get("name")) or ""),
            "Matter Stage": (m.get("matter_stage") or {}).get("name") or "",
            "Responsible Attorney": (m.get("responsible_attorney") or {}).get("name") or "",
            "Trust Account Balance": trust,
            "_cf_map": cf_map
        })
    matter_df = pd.DataFrame(m_rows)

    ocb_df = pd.DataFrame(ocb_rows)          # Client Name + Outstanding Balance
    unbilled_df = pd.DataFrame(unbilled_rows)  # Client Name + Unbilled Amount/Hours

    # Defensive columns
    for df, cols in [
        (matter_df, ["Matter ID","Matter Number","Client Name","Matter Stage","Responsible Attorney","Trust Account Balance","_cf_map"]),
        (ocb_df, ["Client Name","Outstanding Balance"]),
        (unbilled_df, ["Client Name","Unbilled Amount","Unbilled Hours"]),
    ]:
        for c in cols:
            if c not in df.columns:
                df[c] = 0 if ("Amount" in c or "Balance" in c or c.endswith("Hours")) else ({} if c == "_cf_map" else "")

    # -------- Merge EXACTLY like Traffic Light (by Client Name) --------
    combined = (
        matter_df
        .merge(ocb_df, on="Client Name", how="left")
        .merge(unbilled_df, on="Client Name", how="left")
    )

    # Numerics
    for c in ["Trust Account Balance","Outstanding Balance","Unbilled Amount","Unbilled Hours"]:
        combined[c] = pd.to_numeric(combined[c], errors="coerce").fillna(0.0)

    # ✅ Net Trust = Trust – Outstanding – Unbilled (client-level)
    combined["Net Trust Account Balance"] = (
        combined["Trust Account Balance"] - combined["Outstanding Balance"] - combined["Unbilled Amount"]
    ).astype(float)

    # Keep the cycle column for layout parity (0 for case review)
    combined[BILLING_COL] = 0.0

    # Expand custom fields onto columns
    for cf_name in CF_OUTPUT_FIELDS:
        combined[cf_name] = combined["_cf_map"].map(lambda d: (d or {}).get(cf_name, ""))

    # Build output in your expected order
    out_cols = OUTPUT_FIELDS[:]
    for c in out_cols:
        if c not in combined.columns:
            combined[c] = ""
    out = combined[out_cols].copy()
    out = out.sort_values(by="Net Trust Account Balance", ascending=False, kind="mergesort").reset_index(drop=True)
    return out

# ==============================
# Excel writer
# ==============================
def write_excel(df: pd.DataFrame, file_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Case Review"
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([row.get(col, "") for col in df.columns])

    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    net_col = headers.get("Net Trust Account Balance")

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    _ = Font(bold=True)

    for r in range(2, ws.max_row + 1):
        if net_col:
            c = ws.cell(row=r, column=net_col)
            c.number_format = "$#,##0.00"
            try:
                val = float(c.value or 0)
                if val <= 0:
                    c.fill = red_fill
                elif 0 < val < 1000:
                    c.fill = yellow_fill
                else:
                    c.fill = green_fill
            except Exception:
                pass

    ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    table = Table(displayName="CaseReviewTable", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)

    wb.save(file_path)
    print(f"✅ Saved Excel: {file_path}")

# ==============================
# Splitter (SMART)
# ==============================
def _last_name_key(name: str) -> str:
    name = (name or "").strip()
    if not name:
        return ""
    if "," in name:
        last = name.split(",")[0].strip()
    else:
        last = name.split()[-1].strip()
    return last.lower()

def split_and_upload_by_attorney(df: pd.DataFrame, file_date: str):
    """
    Smart behavior:
    - Iterate over every attorney present in the data (by last-name key).
    - Prefer explicit ATTORNEY_* mappings; otherwise derive display/folder as 'Last, First'.
    - Upload to SharePoint; folders are created if missing.
    """
    df = df.copy()
    if "Responsible Attorney" not in df.columns:
        print("⚠️ 'Responsible Attorney' column not found; skipping split.")
        return

    df["_last_key"] = df["Responsible Attorney"].map(_last_name_key)

    try:
        print("Responsible Attorney counts:\n", df["Responsible Attorney"].value_counts())
    except Exception:
        pass

    seen = set()
    for last_key in sorted(k for k in df["_last_key"].unique() if k):
        if last_key in seen:
            continue
        seen.add(last_key)

        atty_df = df[df["_last_key"] == last_key]
        if atty_df.empty:
            continue

        # Prefer explicit display override; else derive from the first full name encountered
        display_name = ATTORNEY_DISPLAY.get(last_key)
        if not display_name:
            sample_name = atty_df["Responsible Attorney"].iloc[0]
            display_name = _to_last_first(sample_name)

        # Prefer explicit folder override; else derive default folder path
        folder = ATTORNEY_FOLDERS.get(last_key) or f"{BASE_ATTY_FOLDER}/{display_name}"

        file_name = f"{file_date}.{display_name} Case Review List.xlsx"
        tmp_path = os.path.join("/tmp", f"case_review_{last_key}.xlsx")

        write_excel(atty_df, tmp_path)
        upload_file(tmp_path, file_name, folder)

        try:
            os.remove(tmp_path)
        except Exception:
            pass

        print(f"✅ Uploaded split file for {display_name} → {folder}/{file_name}")

# ==============================
# Entrypoint helpers
# ==============================
def extract_custom_data_and_build_file() -> tuple[pd.DataFrame, str]:
    df = build_report_dataframe()
    tmp_dir = "/tmp"
    os.makedirs(tmp_dir, exist_ok=True)
    file_path = os.path.join(tmp_dir, "case_review_master.xlsx")
    write_excel(df, file_path)
    return df, file_path

def upload_file(file_path: str, file_name: str, folder_path: str):
    authority = f"https://login.microsoftonline.com/{SHAREPOINT_TENANT_ID}"
    app = msal.ConfidentialClientApplication(
        SHAREPOINT_CLIENT_ID, authority=authority, client_credential=SHAREPOINT_CLIENT_SECRET
    )
    tok = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in tok:
        raise Exception(f"Failed to get Graph token: {tok.get('error_description')}")
    headers = {"Authorization": f"Bearer {tok['access_token']}"}

    # Ensure folder path exists (create if missing)
    parent = ""
    for seg in folder_path.strip("/").split("/"):
        full = f"{parent}/{seg}" if parent else seg
        r = requests.get(
            f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/drives/{SHAREPOINT_DRIVE_ID}/root:/{full}",
            headers=headers
        )
        if r.status_code == 404:
            if parent:
                create_url = (f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/drives/"
                              f"{SHAREPOINT_DRIVE_ID}/root:/{parent}:/children")
            else:
                create_url = (f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/drives/"
                              f"{SHAREPOINT_DRIVE_ID}/root/children")
            requests.post(create_url, headers=headers, json={
                "name": seg, "folder": {}, "@microsoft.graph.conflictBehavior": "replace"
            }).raise_for_status()
        parent = full

    upload_url = (
        f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/drives/{SHAREPOINT_DRIVE_ID}"
        f"/root:/{quote(folder_path + '/' + file_name)}:/content"
    )
    with open(file_path, "rb") as f:
        res = requests.put(upload_url, headers=headers, data=f)
    if res.status_code not in (200, 201):
        raise Exception(f"Upload error: {res.status_code} - {res.text}")
    print(f"✅ Uploaded {file_name} to {folder_path}/")

def main():
    df, file_path = extract_custom_data_and_build_file()
    file_date = datetime.now().strftime("%y%m%d")

    # Master upload
    master_name = f"{file_date}.Seabrook's Case Review List.xlsx"
    upload_file(file_path, master_name, SHAREPOINT_DOC_LIB)

    # Splits (smart)
    split_and_upload_by_attorney(df, file_date)

    # Cleanup
    try:
        os.remove(file_path)
    except Exception:
        pass

if __name__ == "__main__":
    main()
