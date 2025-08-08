# -*- coding: utf-8 -*-
"""
Case Review Report (fixed)
- Correct Net Trust Account Balance (Trust - Outstanding - Unbilled)
- Includes ALL open/pending matters
- Stable joins by IDs; custom field *names* (picklist text) included
- Rate-limit aware requests, 401 refresh, and proper pagination
- Same color formatting & ordering as Traffic Light report
"""

import os
import re
import time
import shutil
import requests
import pandas as pd
from datetime import datetime
from urllib.parse import quote
import msal
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

# ==============================
# Environment / Config
# ==============================
API_VERSION = "4"
CLIO_BASE = os.getenv("CLIO_BASE", "https://app.clio.com").rstrip("/")
CLIO_API = f"{CLIO_BASE}/api/v{API_VERSION}"
CLIO_TOKEN_URL = f"{CLIO_BASE}/oauth/token"

# Clio OAuth (env-backed; uses refresh token)
CLIO_CLIENT_ID = os.getenv("CLIO_CLIENT_ID")
CLIO_CLIENT_SECRET = os.getenv("CLIO_CLIENT_SECRET")
CLIO_REFRESH_TOKEN = os.getenv("CLIO_REFRESH_TOKEN")  # persisted in env
CLIO_ACCESS_TOKEN = os.getenv("CLIO_ACCESS_TOKEN")     # ephemeral
# Backward compat: read old EXPIRES_IN if present; prefer EXPIRES_AT
_CLIO_EXPIRES_AT_ENV = os.getenv("CLIO_EXPIRES_AT") or os.getenv("CLIO_EXPIRES_IN")
try:
    CLIO_EXPIRES_AT = float(_CLIO_EXPIRES_AT_ENV) if _CLIO_EXPIRES_AT_ENV else 0.0
except Exception:
    CLIO_EXPIRES_AT = 0.0
CLIO_REDIRECT_URI = os.getenv("CLIO_REDIRECT_URI")

# SharePoint / Graph
SHAREPOINT_TENANT_ID = os.getenv("GRAPH_TENANT_ID")
SHAREPOINT_CLIENT_ID = os.getenv("GRAPH_CLIENT_ID")
SHAREPOINT_CLIENT_SECRET = os.getenv("GRAPH_CLIENT_SECRET")
SHAREPOINT_SITE_ID = os.getenv("SHAREPOINT_SITE_ID")
SHAREPOINT_DRIVE_ID = os.getenv("SHAREPOINT_DRIVE_ID")
SHAREPOINT_DOC_LIB = os.getenv(
    "SHAREPOINT_DOC_LIB",
    "General Management/Global Case Review Lists"
)

# Report columns (kept as you provided)
OUTPUT_FIELDS = [
    "Matter Number","Client Name","CR ID","Net Trust Account Balance","Matter Stage",
    "Responsible Attorney","Main Paralegal","Client Notes","Initial Client Goals","Initial Strategy",
    "Has strategy changed Describe","Current action Items","Hearings","Deadlines","DV situation description",
    "Custody Visitation","CS Add ons Extracurricular","Spousal Support","PDDs","Discovery",
    "Judgment Trial","Post Judgment","collection efforts"
]

# Rate limiting / pagination
PAGE_LIMIT = 200
GLOBAL_MIN_SLEEP = float(os.getenv("CLIO_GLOBAL_MIN_SLEEP_SEC", "0.25"))

session = requests.Session()
session.headers.update({"Accept": "application/json"})

# ==============================
# Auth Helpers (env-backed)
# ==============================
def _save_tokens_env(tokens: dict):
    """
    Persist ephemeral tokens in environment for the running process.
    Preserve refresh_token if not returned.
    """
    global CLIO_ACCESS_TOKEN, CLIO_EXPIRES_AT, CLIO_REFRESH_TOKEN
    CLIO_ACCESS_TOKEN = tokens.get("access_token", CLIO_ACCESS_TOKEN)
    # Clio commonly does not return refresh_token on refresh; preserve prior
    if tokens.get("refresh_token"):
        CLIO_REFRESH_TOKEN = tokens["refresh_token"]
        os.environ["CLIO_REFRESH_TOKEN"] = CLIO_REFRESH_TOKEN
    # Convert expires_in -> expires_at
    expires_in = tokens.get("expires_in", 0)
    try:
        CLIO_EXPIRES_AT = time.time() + float(expires_in)
    except Exception:
        CLIO_EXPIRES_AT = time.time() + 3000
    os.environ["CLIO_ACCESS_TOKEN"] = CLIO_ACCESS_TOKEN or ""
    os.environ["CLIO_EXPIRES_AT"] = str(CLIO_EXPIRES_AT)

def _refresh_clio_token() -> str:
    global CLIO_ACCESS_TOKEN, CLIO_EXPIRES_AT, CLIO_REFRESH_TOKEN
    if not (CLIO_CLIENT_ID and CLIO_CLIENT_SECRET and CLIO_REFRESH_TOKEN):
        raise RuntimeError("Missing CLIO_CLIENT_ID/CLIO_CLIENT_SECRET/CLIO_REFRESH_TOKEN.")
    resp = session.post(CLIO_TOKEN_URL, data={
        "grant_type": "refresh_token",
        "refresh_token": CLIO_REFRESH_TOKEN,
        "client_id": CLIO_CLIENT_ID,
        "client_secret": CLIO_CLIENT_SECRET
    }, timeout=60)
    if resp.status_code != 200:
        raise RuntimeError(f"Token refresh failed: {resp.status_code} {resp.text[:200]}")
    tokens = resp.json() or {}
    _save_tokens_env(tokens)
    return CLIO_ACCESS_TOKEN

def _get_clio_token() -> str:
    global CLIO_ACCESS_TOKEN, CLIO_EXPIRES_AT
    now = time.time()
    if not CLIO_ACCESS_TOKEN or now >= (CLIO_EXPIRES_AT or 0):
        return _refresh_clio_token()
    return CLIO_ACCESS_TOKEN

def _ensure_auth_header():
    token = _get_clio_token()
    session.headers["Authorization"] = f"Bearer {token}"

# ==============================
# Backoff / Request Wrapper
# ==============================
def _sleep_with_floor(start_ts: float, retry_after: int | float | None = None):
    elapsed = time.time() - start_ts
    base_wait = max(0, GLOBAL_MIN_SLEEP - elapsed)
    wait = max(base_wait, float(retry_after or 0))
    if wait > 0:
        time.sleep(wait)

def _request(method: str, url: str, **kwargs) -> requests.Response:
    _ensure_auth_header()
    max_tries = kwargs.pop("max_tries", 7)
    backoff = 1
    for _ in range(max_tries):
        t0 = time.time()
        resp = session.request(method, url, timeout=60, **kwargs)

        # Token expiry / auth issues -> refresh once per attempt
        if resp.status_code == 401:
            try:
                _refresh_clio_token()
            finally:
                _sleep_with_floor(t0, retry_after=1)
            continue

        # Rate limit
        if resp.status_code == 429:
            ra = 30
            if resp.headers.get("Retry-After"):
                try: ra = int(resp.headers["Retry-After"])
                except ValueError: ra = 30
            else:
                try:
                    msg = (resp.json() or {}).get("error", {}).get("message", "")
                    m = re.search(r"Retry in (\d+)", msg)
                    if m: ra = int(m.group(1))
                except Exception:
                    pass
            print(f"[429] Rate limited. Waiting {ra}s …")
            _sleep_with_floor(t0, retry_after=ra)
            continue

        # Server errors
        if 500 <= resp.status_code < 600:
            print(f"[{resp.status_code}] Server error. Retrying in {backoff}s …")
            _sleep_with_floor(t0, retry_after=backoff)
            backoff = min(backoff * 2, 60)
            continue

        _sleep_with_floor(t0)
        return resp
    return resp  # last response (likely error)

# ==============================
# Proper Clio pagination (meta.paging.next)
# ==============================
def paginate(url: str, params: dict | None = None):
    params = dict(params or {})
    params.setdefault("limit", PAGE_LIMIT)
    params.setdefault("order", "id(asc)")  # stable cursor pagination

    next_url = url
    next_params = params
    all_rows = []

    while True:
        resp = _request("GET", next_url, params=next_params)
        print(f"GET {next_url} params={next_params} → {resp.status_code}")
        if resp.status_code != 200:
            print(f"Failed: {resp.status_code} {resp.text[:200]}")
            break

        body = resp.json() or {}
        rows = body.get("data", [])
        if isinstance(rows, list):
            all_rows.extend([r for r in rows if isinstance(r, dict)])

        paging = (body.get("meta") or {}).get("paging") or {}
        next_link = paging.get("next")

        if next_link:
            next_url = next_link
            next_params = None
            continue

        # Fallback: stop when short page and no 'next'
        if len(rows) < params["limit"]:
            break
        else:
            break  # avoid accidental infinite loop

    return all_rows

# ==============================
# Clio data fetchers
# ==============================
def fetch_open_matters_with_cf():
    """
    Matters with account balances and required headers + custom fields.
    """
    url = f"{CLIO_API}/matters.json"
    fields = (
        "id,display_number,number,"
        "client{id,name},"
        "matter_stage{name},"
        "responsible_attorney{name},"
        "account_balances{balance},"
        "custom_field_values{id,field_name,field_type,value,picklist_option}"
    )
    params = {"status": "open,pending", "fields": fields}
    return paginate(url, params)

def fetch_outstanding_client_balances():
    """
    Outstanding AR per client (contact).
    """
    url = f"{CLIO_API}/outstanding_client_balances.json"
    params = {"fields": "contact{id,name},total_outstanding_balance"}
    return paginate(url, params)

def fetch_billable_matters():
    """
    Unbilled amounts/hours per *matter*.
    """
    url = f"{CLIO_API}/billable_matters.json"
    fields = "id,display_number,client{id,name},unbilled_amount,unbilled_hours"
    params = {"fields": fields}
    return paginate(url, params)

# ==============================
# SharePoint (unchanged)
# ==============================
def ensure_folder(path: str, headers: dict):
    segments = path.strip("/").split("/")
    parent = ""
    for seg in segments:
        full = f"{parent}/{seg}" if parent else seg
        url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/drives/{SHAREPOINT_DRIVE_ID}/root:/{full}"
        r = requests.get(url, headers=headers)
        if r.status_code == 404:
            if parent:
                create_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/drives/{SHAREPOINT_DRIVE_ID}/root:/{parent}:/children"
            else:
                create_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/drives/{SHAREPOINT_DRIVE_ID}/root/children"
            create = requests.post(create_url, headers=headers, json={
                "name": seg,
                "folder": {},
                "@microsoft.graph.conflictBehavior": "replace"
            })
            create.raise_for_status()
        parent = full

def upload_file(file_path: str, file_name: str, folder_path: str):
    authority = f"https://login.microsoftonline.com/{SHAREPOINT_TENANT_ID}"
    app = msal.ConfidentialClientApplication(
        SHAREPOINT_CLIENT_ID,
        authority=authority,
        client_credential=SHAREPOINT_CLIENT_SECRET
    )
    tok = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in tok:
        raise Exception(f"Failed to get Graph token: {tok.get('error_description')}")

    headers = {"Authorization": f"Bearer {tok['access_token']}", "Content-Type": "application/json"}
    ensure_folder(folder_path, headers)

    encoded_path = quote(f"{folder_path}/{file_name}")
    upload_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/drives/{SHAREPOINT_DRIVE_ID}/root:/{encoded_path}:/content"

    with open(file_path, "rb") as f:
        res = requests.put(upload_url, headers={"Authorization": f"Bearer {tok['access_token']}"}, data=f)
    if res.status_code not in (200, 201):
        raise Exception(f"Upload error: {res.status_code} - {res.text}")
    print(f"✅ Uploaded {file_name} to SharePoint at {folder_path}/")

# ==============================
# Report generation
# ==============================
def build_report_dataframe() -> pd.DataFrame:
    """
    Collect data from Clio and return a DataFrame with OUTPUT_FIELDS
    and computed Net Trust Account Balance, sorted descending.
    """
    # Base matters (includes balances + custom fields)
    matters = fetch_open_matters_with_cf()
    # Outstanding balances per client
    ocb = fetch_outstanding_client_balances()
    # Unbilled per matter
    billable = fetch_billable_matters()

    # --- Build frames
    # matters base
    m_rows = []
    for m in matters:
        acc_bal = 0.0
        for b in (m.get("account_balances") or []):
            try:
                acc_bal += float((b or {}).get("balance", 0) or 0)
            except Exception:
                pass

        m_rows.append({
            "Matter ID": m.get("id"),
            "Matter Number": m.get("display_number") or m.get("number") or "",
            "Client ID": (m.get("client") or {}).get("id"),
            "Client Name": (m.get("client") or {}).get("name") or "",
            "Matter Stage": (m.get("matter_stage") or {}).get("name") or "",
            "Responsible Attorney": (m.get("responsible_attorney") or {}).get("name") or "",
            "Trust Account Balance": acc_bal,
            # stash raw custom fields to extract named ones below
            "_custom_field_values": m.get("custom_field_values") or []
        })

    matter_df = pd.DataFrame(m_rows)

    # outstanding by client
    ocb_rows = []
    for r in ocb:
        ocb_rows.append({
            "Client ID": (r.get("contact") or {}).get("id"),
            "Outstanding Balance": r.get("total_outstanding_balance", 0) or 0
        })
    ocb_df = pd.DataFrame(ocb_rows)

    # unbilled per matter
    b_rows = []
    for bm in billable:
        b_rows.append({
            "Matter ID": bm.get("id"),
            "Matter Number": bm.get("display_number") or "",
            "Client ID": (bm.get("client") or {}).get("id"),
            "Unbilled Amount": bm.get("unbilled_amount", 0) or 0,
            "Unbilled Hours": bm.get("unbilled_hours", 0) or 0
        })
    billable_df = pd.DataFrame(b_rows)

    # Ensure base columns exist even if empty results to avoid KeyError
    for df, cols in [
        (matter_df, ["Matter ID","Matter Number","Client ID","Client Name","Matter Stage","Responsible Attorney","Trust Account Balance","_custom_field_values"]),
        (ocb_df, ["Client ID","Outstanding Balance"]),
        (billable_df, ["Matter ID","Matter Number","Client ID","Unbilled Amount","Unbilled Hours"]),
    ]:
        for c in cols:
            if c not in df.columns:
                df[c] = 0 if "Amount" in c or c.endswith("Balance") else ""

    # --- Merge (base = matters; include ALL open/pending matters)
    combined = (
        matter_df
        .merge(ocb_df, on="Client ID", how="left")
        .merge(billable_df[["Matter ID","Unbilled Amount","Unbilled Hours"]], on="Matter ID", how="left")
    )

    # Fill missing numeric fields
    for c in ["Trust Account Balance","Outstanding Balance","Unbilled Amount","Unbilled Hours"]:
        if c in combined.columns:
            combined[c] = pd.to_numeric(combined[c], errors="coerce").fillna(0.0)

    # --- Compute Net Trust Account Balance (same as Traffic Light report)
    combined["Net Trust Account Balance"] = (
        combined["Trust Account Balance"] - combined["Outstanding Balance"] - combined["Unbilled Amount"]
    ).astype(float)

    # --- Extract custom fields into named columns (picklist -> option text)
    def extract_cf_map(cf_list):
        out = {}
        for cf in (cf_list or []):
            if not isinstance(cf, dict):
                continue
            fname = cf.get("field_name")
            if not fname:
                continue
            val = cf.get("value")
            if (not val) and isinstance(cf.get("picklist_option"), dict):
                val = cf["picklist_option"].get("option", "")
            out[fname] = val or ""
        return out

    combined["_cf_map"] = combined["_custom_field_values"].apply(extract_cf_map)

    # --- Build output rows in the exact OUTPUT_FIELDS order
    rows = []
    for _, r in combined.iterrows():
        cf = r["_cf_map"]

        row = {
            "Matter Number": r.get("Matter Number", ""),
            "Client Name": r.get("Client Name", ""),
            "CR ID": cf.get("CR ID", ""),
            "Net Trust Account Balance": float(r.get("Net Trust Account Balance", 0.0)),
            "Matter Stage": r.get("Matter Stage", ""),
            "Responsible Attorney": r.get("Responsible Attorney", ""),
            "Main Paralegal": cf.get("Main Paralegal", ""),                 # picklist option text
            "Client Notes": cf.get("Client Notes", ""),
            "Initial Client Goals": cf.get("Initial Client Goals", ""),
            "Initial Strategy": cf.get("Initial Strategy", ""),
            "Has strategy changed Describe": cf.get("Has strategy changed Describe", ""),
            "Current action Items": cf.get("Current action Items", ""),
            "Hearings": cf.get("Hearings", ""),
            "Deadlines": cf.get("Deadlines", ""),
            "DV situation description": cf.get("DV situation description", ""),
            "Custody Visitation": cf.get("Custody Visitation", ""),
            "CS Add ons Extracurricular": cf.get("CS Add ons Extracurricular", ""),
            "Spousal Support": cf.get("Spousal Support", ""),
            "PDDs": cf.get("PDDs", ""),
            "Discovery": cf.get("Discovery", ""),
            "Judgment Trial": cf.get("Judgment Trial", ""),
            "Post Judgment": cf.get("Post Judgment", ""),
            "collection efforts": cf.get("collection efforts", "")
        }
        rows.append(row)

    df = pd.DataFrame(rows, columns=OUTPUT_FIELDS)

    # --- Sort by Net Trust Account Balance (desc)
    df = df.sort_values(by="Net Trust Account Balance", ascending=False, kind="mergesort").reset_index(drop=True)
    return df

def write_excel(df: pd.DataFrame, file_path: str):
    """
    Writes the DataFrame to Excel with:
    - currency formatting for Net Trust
    - traffic-light conditional fills (same thresholds)
    - table styling
    """
    # Write with openpyxl to control formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Case Review"

    # Headers
    ws.append(list(df.columns))

    # Rows
    for _, row in df.iterrows():
        ws.append([row.get(col, "") for col in df.columns])

    # Locate Net Trust column index
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    net_col = headers.get("Net Trust Account Balance", None)

    # Styles (same as Traffic Light)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    bold_font = Font(bold=True)

    # Format data rows
    for r in range(2, ws.max_row + 1):
        if net_col:
            c = ws.cell(row=r, column=net_col)
            # currency number format
            c.number_format = "$#,##0.00"
            # traffic light fill
            try:
                val = float(c.value or 0)
                if val <= 0:
                    c.fill = red_fill
                elif 0 < val < 1000:
                    c.fill = yellow_fill
                else:
                    c.fill = green_fill
            except Exception:
                pass

    # Add table styling
    ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    table = Table(displayName="CaseReviewTable", ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save file
    wb.save(file_path)
    print(f"✅ Saved Excel: {file_path}")

def extract_custom_data_and_build_file() -> str:
    df = build_report_dataframe()

    # Save to a temp file (Render)
    tmp_dir = "/tmp"
    os.makedirs(tmp_dir, exist_ok=True)
    file_path = os.path.join(tmp_dir, "case_review.xlsx")

    write_excel(df, file_path)
    return file_path

def main():
    # Build report and upload
    file_path = extract_custom_data_and_build_file()
    file_date = datetime.now().strftime("%y%m%d")
    file_name = f"{file_date}.Seabrook's Case Review List.xlsx"
    upload_file(file_path, file_name, SHAREPOINT_DOC_LIB)

    # Cleanup temp
    try:
        os.remove(file_path)
    except Exception:
        pass

if __name__ == "__main__":
    main()


if __name__ == '__main__':
    main()
